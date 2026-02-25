"""
Generate Excel templates for import using openpyxl (no pandas dependency).
Run: python create_templates.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

templates_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app', 'excel_templates')
os.makedirs(templates_dir, exist_ok=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(bold=True, size=11, color='FFFFFF')
example_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')


def style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


def style_data(ws, row, col_count, fill=None):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
        if fill:
            cell.fill = fill


# ============================================================
# 1. Teachers template
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = 'Учителя'

headers = ['ФИО', 'Краткое имя', 'Email', 'Телефон']
widths = [40, 20, 30, 25]

for i, (h, w) in enumerate(zip(headers, widths), 1):
    ws.cell(row=1, column=i, value=h)
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

style_header(ws, 1, len(headers))

examples = [
    ['Иванова Мария Петровна', 'Иванова М.П.', 'ivanova@school.ru', '+7 (999) 123-45-67'],
    ['Петров Сергей Николаевич', 'Петров С.Н.', 'petrov@school.ru', '+7 (999) 234-56-78'],
    ['Сидорова Анна Ивановна', 'Сидорова А.И.', 'sidorova@school.ru', '+7 (999) 345-67-89'],
]

for row_idx, row_data in enumerate(examples, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=val)
    style_data(ws, row_idx, len(headers), fill=example_fill)

ws.cell(row=6, column=1, value='* Обязательное поле — только ФИО. Остальные столбцы необязательны.')
ws.cell(row=6, column=1).font = Font(italic=True, color='888888', size=10)
ws.cell(row=7, column=1, value='* Удалите примеры выше перед импортом и заполните своими данными.')
ws.cell(row=7, column=1).font = Font(italic=True, color='888888', size=10)

path1 = os.path.join(templates_dir, 'teachers_template.xlsx')
wb.save(path1)
print(f'Created: {path1}')


# ============================================================
# 2. Curriculum template — Elementary (1-4 classes)
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = 'Учебный план (нач.)'

subjects = [
    'Русский язык', 'Математика', 'Литературное чтение',
    'Окружающий мир', 'Английский язык', 'Музыка',
    'ИЗО', 'Технология', 'Физкультура',
]

classes = ['1А', '1Б', '2А', '2Б', '3А', '3Б', '4А', '4Б']

hours = [
    [5, 4, 4, 2, 0, 1, 1, 1, 3],  # 1А
    [5, 4, 4, 2, 0, 1, 1, 1, 3],  # 1Б
    [5, 4, 4, 2, 2, 1, 1, 1, 3],  # 2А
    [5, 4, 4, 2, 2, 1, 1, 1, 3],  # 2Б
    [4, 4, 3, 2, 2, 1, 1, 1, 3],  # 3А
    [4, 4, 3, 2, 2, 1, 1, 1, 3],  # 3Б
    [4, 4, 3, 2, 2, 1, 1, 1, 3],  # 4А
    [4, 4, 3, 2, 2, 1, 1, 1, 3],  # 4Б
]

ws.cell(row=1, column=1, value='Класс')
for j, subj in enumerate(subjects, 2):
    ws.cell(row=1, column=j, value=subj)

ws.column_dimensions['A'].width = 10
for j in range(2, len(subjects) + 2):
    ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = 22

style_header(ws, 1, len(subjects) + 1)

for i, (cls, hrs) in enumerate(zip(classes, hours), 2):
    ws.cell(row=i, column=1, value=cls)
    ws.cell(row=i, column=1).font = Font(bold=True)
    ws.cell(row=i, column=1).alignment = Alignment(horizontal='center')
    for j, h in enumerate(hrs, 2):
        ws.cell(row=i, column=j, value=h)
        ws.cell(row=i, column=j).alignment = Alignment(horizontal='center')
    style_data(ws, i, len(subjects) + 1, fill=example_fill if i % 2 == 0 else None)

note_row = len(classes) + 3
ws.cell(row=note_row, column=1, value='* Первый столбец — название класса (например 1А, 2Б).')
ws.cell(row=note_row, column=1).font = Font(italic=True, color='888888', size=10)
ws.cell(row=note_row + 1, column=1, value='* Остальные столбцы — предметы. В ячейках — часы в неделю. 0 = предмет не ведётся.')
ws.cell(row=note_row + 1, column=1).font = Font(italic=True, color='888888', size=10)
ws.cell(row=note_row + 2, column=1, value='* Удалите примеры и заполните своими данными перед импортом.')
ws.cell(row=note_row + 2, column=1).font = Font(italic=True, color='888888', size=10)

path2 = os.path.join(templates_dir, 'curriculum_elementary_template.xlsx')
wb.save(path2)
print(f'Created: {path2}')


# ============================================================
# 3. Curriculum template — Secondary (5-11 classes)
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = 'Учебный план (ст.)'

subjects_sec = [
    'Русский язык', 'Литература', 'Математика', 'Алгебра', 'Геометрия',
    'Английский язык', 'История', 'Обществознание', 'География',
    'Биология', 'Физика', 'Химия', 'Информатика',
    'Музыка', 'ИЗО', 'Технология', 'Физкультура', 'ОБЖ',
]

classes_sec = ['5А', '5Б', '6А', '6Б', '7А', '7Б', '8А', '8Б', '9А', '9Б', '10А', '11А']

hours_sec = [
    [5, 3, 5, 0, 0, 3, 2, 0, 1, 1, 0, 0, 0, 1, 1, 2, 3, 0],  # 5А
    [5, 3, 5, 0, 0, 3, 2, 0, 1, 1, 0, 0, 0, 1, 1, 2, 3, 0],  # 5Б
    [6, 3, 5, 0, 0, 3, 2, 1, 1, 1, 0, 0, 0, 1, 1, 2, 3, 0],  # 6А
    [6, 3, 5, 0, 0, 3, 2, 1, 1, 1, 0, 0, 0, 1, 1, 2, 3, 0],  # 6Б
    [4, 2, 0, 3, 2, 3, 2, 1, 2, 1, 2, 0, 1, 1, 1, 2, 3, 0],  # 7А
    [4, 2, 0, 3, 2, 3, 2, 1, 2, 1, 2, 0, 1, 1, 1, 2, 3, 0],  # 7Б
    [3, 2, 0, 3, 2, 3, 2, 1, 2, 2, 2, 2, 1, 0, 0, 1, 3, 1],  # 8А
    [3, 2, 0, 3, 2, 3, 2, 1, 2, 2, 2, 2, 1, 0, 0, 1, 3, 1],  # 8Б
    [3, 3, 0, 3, 2, 3, 3, 1, 2, 2, 3, 2, 1, 0, 0, 0, 3, 1],  # 9А
    [3, 3, 0, 3, 2, 3, 3, 1, 2, 2, 3, 2, 1, 0, 0, 0, 3, 1],  # 9Б
    [2, 3, 0, 3, 2, 3, 2, 2, 1, 1, 3, 2, 1, 0, 0, 0, 3, 1],  # 10А
    [2, 3, 0, 3, 2, 3, 2, 2, 1, 1, 3, 2, 1, 0, 0, 0, 3, 1],  # 11А
]

ws.cell(row=1, column=1, value='Класс')
for j, subj in enumerate(subjects_sec, 2):
    ws.cell(row=1, column=j, value=subj)

ws.column_dimensions['A'].width = 10
for j in range(2, len(subjects_sec) + 2):
    ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = 18

style_header(ws, 1, len(subjects_sec) + 1)

for i, (cls, hrs) in enumerate(zip(classes_sec, hours_sec), 2):
    ws.cell(row=i, column=1, value=cls)
    ws.cell(row=i, column=1).font = Font(bold=True)
    ws.cell(row=i, column=1).alignment = Alignment(horizontal='center')
    for j, h in enumerate(hrs, 2):
        ws.cell(row=i, column=j, value=h)
        ws.cell(row=i, column=j).alignment = Alignment(horizontal='center')
    style_data(ws, i, len(subjects_sec) + 1, fill=example_fill if i % 2 == 0 else None)

note_row = len(classes_sec) + 3
ws.cell(row=note_row, column=1, value='* Первый столбец — название класса (например 5А, 10Б).')
ws.cell(row=note_row, column=1).font = Font(italic=True, color='888888', size=10)
ws.cell(row=note_row + 1, column=1, value='* Остальные столбцы — предметы. В ячейках — часы в неделю. 0 = предмет не ведётся.')
ws.cell(row=note_row + 1, column=1).font = Font(italic=True, color='888888', size=10)
ws.cell(row=note_row + 2, column=1, value='* Удалите примеры и заполните своими данными перед импортом.')
ws.cell(row=note_row + 2, column=1).font = Font(italic=True, color='888888', size=10)

path3 = os.path.join(templates_dir, 'curriculum_secondary_template.xlsx')
wb.save(path3)
print(f'Created: {path3}')


# ============================================================
# 4. Classrooms template
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = 'Кабинеты'

headers = ['Номер', 'Название', 'Вместимость классов', 'Этаж', 'Корпус']
widths = [12, 30, 22, 10, 15]

for i, (h, w) in enumerate(zip(headers, widths), 1):
    ws.cell(row=1, column=i, value=h)
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

style_header(ws, 1, len(headers))

examples = [
    ['101', 'Кабинет математики', 1, 1, 'Основной'],
    ['205', 'Кабинет русского языка', 1, 2, 'Основной'],
    ['301', 'Кабинет физики', 1, 3, 'Основной'],
    ['Спортзал', 'Спортивный зал', 3, 1, 'Основной'],
]

for row_idx, row_data in enumerate(examples, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=val)
    style_data(ws, row_idx, len(headers), fill=example_fill)

note_row = len(examples) + 3
ws.cell(row=note_row, column=1, value='* Обязательное поле — только Номер. Остальные столбцы необязательны.')
ws.cell(row=note_row, column=1).font = Font(italic=True, color='888888', size=10)

path4 = os.path.join(templates_dir, 'classrooms_template.xlsx')
wb.save(path4)
print(f'Created: {path4}')

print('\nAll templates created successfully!')
print(f'Location: {os.path.abspath(templates_dir)}')
