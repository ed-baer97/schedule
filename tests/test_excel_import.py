"""Import of per-subject teacher × class hour matrices."""
from __future__ import annotations

from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import delete, select

from app.models import SchoolClass, Subject, Teacher, TeachingAssignment
from app.services.excel_import import ExcelImporter
from backend.deps import SessionLocal
from backend.main import app
from tests.conftest import TEST_SCHOOL_ID

client = TestClient(app)


@pytest.fixture(autouse=True)
def _clear_import_data() -> None:
    with SessionLocal() as session:
        for model in (TeachingAssignment, SchoolClass, Subject, Teacher):
            session.execute(delete(model))
        session.commit()


def _write_hours_xlsx(
    path: Path,
    *,
    header: list[str],
    rows: list[list[object]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _write_school_hours_xlsx(
    path: Path,
    *,
    sheet: str,
    subject: str,
    classes: list[str],
    rows: list[tuple[str, dict[str, int]]],
    year: str = "2026-2027",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws["B1"] = subject
    ws["V1"] = year
    headers = ["№", "ФИО", *classes, "итого"]
    for col, title in enumerate(headers, start=1):
        ws.cell(2, col, title)
    for index, (name, hours_by_class) in enumerate(rows, start=1):
        excel_row = index + 2
        ws.cell(excel_row, 1, index)
        ws.cell(excel_row, 2, name)
        total = 0
        for offset, class_name in enumerate(classes):
            hours = hours_by_class.get(class_name)
            if hours:
                ws.cell(excel_row, 3 + offset, hours)
                total += hours
        ws.cell(excel_row, 3 + len(classes), total)
    wb.save(path)


def test_subject_hours_creates_teachers_classes_subgroups(tmp_path: Path) -> None:
    path = tmp_path / "Английский язык.xlsx"
    _write_hours_xlsx(
        path,
        header=["Учитель", "5А", "5Б", "1А"],
        rows=[
            ["Иванова М.П.", 2, 2, 0],
            ["Петров С.Н.", 2, "", 3],
        ],
    )

    with SessionLocal() as session:
        result = ExcelImporter(session, school_id=TEST_SCHOOL_ID).import_subject_hours(
            str(path)
        )

    assert result["subject"] == "Английский язык"
    assert result["teachers_created"] == 2
    assert result["classes_created"] == 3
    assert result["subgroup_classes"] == 1

    with SessionLocal() as session:
        teachers = list(session.scalars(select(Teacher)).all())
        assert {t.full_name for t in teachers} == {"Иванова М.П.", "Петров С.Н."}

        classes = {c.name: c for c in session.scalars(select(SchoolClass)).all()}
        assert classes["5А"].school_level == "secondary"
        assert classes["1А"].school_level == "elementary"

        assignments = list(session.scalars(select(TeachingAssignment)).all())
        # 5А: two teachers → subgroups; 5Б: one; 1А: one (zero hours skipped for Иванова)
        assert len(assignments) == 4
        five_a = [a for a in assignments if a.school_class.name == "5А"]
        assert sorted(a.group_number for a in five_a) == [1, 2]
        five_b = [a for a in assignments if a.school_class.name == "5Б"]
        assert len(five_b) == 1
        assert five_b[0].group_number is None
        assert five_b[0].hours_per_week == 2


def test_same_teacher_in_two_subject_files(tmp_path: Path) -> None:
    math_path = tmp_path / "Математика.xlsx"
    phys_path = tmp_path / "Физика.xlsx"
    _write_hours_xlsx(
        math_path,
        header=["Учитель", "7А"],
        rows=[["Сидоров А.А.", 5]],
    )
    _write_hours_xlsx(
        phys_path,
        header=["Учитель", "7А"],
        rows=[["Сидоров А.А.", 2]],
    )

    with SessionLocal() as session:
        importer = ExcelImporter(session, school_id=TEST_SCHOOL_ID)
        importer.import_subject_hours(str(math_path))
        importer.import_subject_hours(str(phys_path))

    with SessionLocal() as session:
        teachers = list(session.scalars(select(Teacher)).all())
        assert len(teachers) == 1
        subjects = {s.name for s in session.scalars(select(Subject)).all()}
        assert subjects == {"Математика", "Физика"}
        assignments = list(session.scalars(select(TeachingAssignment)).all())
        assert len(assignments) == 2
        assert {a.hours_per_week for a in assignments} == {5, 2}
        assert all(a.teacher_id == teachers[0].id for a in assignments)


def test_api_subject_hours_uses_filename(tmp_path: Path) -> None:
    path = tmp_path / "История.xlsx"
    _write_hours_xlsx(
        path,
        header=["Учитель", "8А"],
        rows=[["Козлова Е.В.", 2]],
    )
    with path.open("rb") as fh:
        r = client.post(
            "/api/import/subject-hours",
            files=[("files", ("История.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
        )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["files"][0]["subject"] == "История"
    assert body["files"][0]["assignments_created"] == 1
    assert "История" in body["message"]


def test_school_workbook_reads_fio_skips_totals(tmp_path: Path) -> None:
    path = tmp_path / "Инф.xlsx"
    _write_school_hours_xlsx(
        path,
        sheet="Информатика",
        subject="Информатика",
        classes=["1А", "5А", "5Б"],
        rows=[
            ("Рскалиева Г.", {"1А": 1}),
            ("Баер Э.В.", {"5А": 2, "5Б": 2}),
            ("Өтепбергенов М.М.", {"5А": 2, "5Б": 2}),
        ],
    )

    with SessionLocal() as session:
        result = ExcelImporter(session, school_id=TEST_SCHOOL_ID).import_subject_hours(
            str(path), filename="Инф.xlsx"
        )

    assert result["subject"] == "Информатика"
    assert result["teachers_created"] == 3
    assert result["classes_created"] == 3
    assert result["subgroup_classes"] == 2

    with SessionLocal() as session:
        names = {t.full_name for t in session.scalars(select(Teacher)).all()}
        assert names == {"Рскалиева Г.", "Баер Э.В.", "Өтепбергенов М.М."}
        class_names = {c.name for c in session.scalars(select(SchoolClass)).all()}
        assert class_names == {"1А", "5А", "5Б"}
        assert "итого" not in class_names
        assignments = list(session.scalars(select(TeachingAssignment)).all())
        assert len(assignments) == 5
        one_a = [a for a in assignments if a.school_class.name == "1А"]
        assert len(one_a) == 1
        assert one_a[0].group_number is None
        five_a = [a for a in assignments if a.school_class.name == "5А"]
        assert sorted(a.group_number for a in five_a) == [1, 2]


def test_same_fio_in_second_subject_is_same_teacher(tmp_path: Path) -> None:
    informatics = tmp_path / "Инф.xlsx"
    math = tmp_path / "Математика.xlsx"
    _write_school_hours_xlsx(
        informatics,
        sheet="Информатика",
        subject="Информатика",
        classes=["5А"],
        rows=[("Баер Э.В.", {"5А": 2})],
    )
    _write_school_hours_xlsx(
        math,
        sheet="Математика",
        subject="Математика",
        classes=["5А"],
        rows=[("Баер Э.В.", {"5А": 5})],
    )

    with SessionLocal() as session:
        importer = ExcelImporter(session, school_id=TEST_SCHOOL_ID)
        importer.import_subject_hours(str(informatics), filename="Инф.xlsx")
        importer.import_subject_hours(str(math), filename="Математика.xlsx")

    with SessionLocal() as session:
        teachers = list(session.scalars(select(Teacher)).all())
        assert len(teachers) == 1
        assert teachers[0].full_name == "Баер Э.В."
        subjects = {s.name for s in session.scalars(select(Subject)).all()}
        assert subjects == {"Информатика", "Математика"}
        assignments = list(session.scalars(select(TeachingAssignment)).all())
        assert len(assignments) == 2
        assert {a.hours_per_week for a in assignments} == {2, 5}


def test_api_school_workbook_uses_subject_from_file(tmp_path: Path) -> None:
    path = tmp_path / "Инф.xlsx"
    _write_school_hours_xlsx(
        path,
        sheet="Информатика",
        subject="Информатика",
        classes=["8А"],
        rows=[("Козлова Е.В.", {"8А": 2})],
    )
    with path.open("rb") as fh:
        r = client.post(
            "/api/import/subject-hours",
            files=[
                (
                    "files",
                    (
                        "Инф.xlsx",
                        fh,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ),
                )
            ],
        )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["files"][0]["subject"] == "Информатика"
    assert body["files"][0]["teachers_created"] == 1
    assert body["files"][0]["assignments_created"] == 1


_REAL_INFORMATICS = Path(r"c:\Users\eduar\Downloads\Инф.xlsx")


@pytest.mark.skipif(not _REAL_INFORMATICS.exists(), reason="sample workbook missing")
def test_real_informatics_workbook() -> None:
    with SessionLocal() as session:
        result = ExcelImporter(session, school_id=TEST_SCHOOL_ID).import_subject_hours(
            str(_REAL_INFORMATICS), filename="Инф.xlsx"
        )

    assert result["subject"] == "Информатика"
    assert result["teachers_created"] == 9
    assert result["classes_created"] == 56

    with SessionLocal() as session:
        class_names = {c.name for c in session.scalars(select(SchoolClass)).all()}
        assert "итого" not in class_names
        assert "ФИО" not in {t.full_name for t in session.scalars(select(Teacher)).all()}
        names = {t.full_name for t in session.scalars(select(Teacher)).all()}
        assert "Баер Э.В." in names
        assert "Өтепбергенов М.М." in names
        five_a = [
            a
            for a in session.scalars(select(TeachingAssignment)).all()
            if a.school_class.name == "5А"
        ]
        assert len(five_a) == 2
        assert sorted(a.group_number for a in five_a) == [1, 2]
