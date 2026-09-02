"""Restore schedule cells from a timetable Excel export."""
from __future__ import annotations

from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import delete, select

from app.models import (
    Classroom,
    Job,
    ScheduleCell,
    ScheduleSettings,
    SchoolClass,
    Shift,
    ShiftLessonTime,
    Subject,
    Teacher,
    TeachingAssignment,
    classroom_subjects,
)
from app.services.report_service import ReportService
from app.services.schedule_excel import parse_cell_entries, parse_schedule_workbook
from backend.deps import SessionLocal
from backend.main import app
from tests.conftest import TEST_SCHOOL_ID

client = TestClient(app)


@pytest.fixture(autouse=True)
def _clear_db() -> None:
    with SessionLocal() as session:
        for model in (
            ScheduleCell,
            TeachingAssignment,
            ShiftLessonTime,
            ScheduleSettings,
            Job,
            SchoolClass,
            Shift,
            classroom_subjects,
            Classroom,
            Subject,
            Teacher,
        ):
            session.execute(delete(model))
        session.commit()


def test_parse_cell_entries_slash_and_group() -> None:
    entries = parse_cell_entries(
        "08:25–09:10\nМатематика · гр.2 / каб. 12 / Сидоров С.С."
    )
    assert len(entries) == 1
    assert entries[0].subject == "Математика"
    assert entries[0].group_number == 2
    assert entries[0].classroom == "12"
    assert entries[0].teacher == "Сидоров С.С."


def test_parse_cell_entries_two_groups() -> None:
    text = (
        "Английский язык · гр.1 / каб. 3 / Иванова А.А.\n"
        "Английский язык · гр.2 / каб. 4 / Петрова Б.Б."
    )
    entries = parse_cell_entries(text)
    assert [e.group_number for e in entries] == [1, 2]
    assert [e.teacher for e in entries] == ["Иванова А.А.", "Петрова Б.Б."]


def test_parse_cell_entries_teacher_format() -> None:
    entries = parse_cell_entries("1Б / Математика / каб. 12")
    assert len(entries) == 1
    assert entries[0].class_name == "1Б"
    assert entries[0].subject == "Математика"
    assert entries[0].classroom == "12"
    assert entries[0].teacher is None


def test_parse_cell_entries_skips_class_hour() -> None:
    assert parse_cell_entries("Классный час") == []
    assert parse_cell_entries("") == []


def test_parse_cell_entries_legacy_block() -> None:
    entries = parse_cell_entries("Математика\nкаб. 12\nСидоров С.С.")
    assert len(entries) == 1
    assert entries[0].subject == "Математика"
    assert entries[0].classroom == "12"
    assert entries[0].teacher == "Сидоров С.С."


def _seed_one_lesson(**kwargs) -> tuple[int, int]:
    group_number = kwargs.get("group_number")
    with SessionLocal() as session:
        shift = Shift(
            school_id=TEST_SCHOOL_ID,
            name="1 смена",
            school_level="elementary",
            start_lesson=1,
            lessons_count=5,
            working_days=5,
            max_lessons_per_day=5,
        )
        subject = Subject(school_id=TEST_SCHOOL_ID, name="Математика")
        teacher = Teacher(school_id=TEST_SCHOOL_ID, full_name="Сидоров С.С.")
        cls = SchoolClass(
            school_id=TEST_SCHOOL_ID, name="1Б", grade=1, school_level="elementary"
        )
        room = Classroom(school_id=TEST_SCHOOL_ID, number="12")
        session.add_all([shift, subject, teacher, cls, room])
        session.flush()
        cls.shift_id = shift.id
        assignment = TeachingAssignment(
            school_id=TEST_SCHOOL_ID,
            subject_id=subject.id,
            teacher_id=teacher.id,
            class_id=cls.id,
            hours_per_week=4,
            group_number=group_number,
        )
        session.add(assignment)
        session.flush()
        session.add(
            ScheduleCell(
                school_id=TEST_SCHOOL_ID,
                class_id=cls.id,
                day_of_week=1,
                lesson_number=1,
                assignment_id=assignment.id,
                classroom_id=room.id,
            )
        )
        ids = (cls.id, assignment.id)
        session.commit()
        return ids


def test_export_all_roundtrip_restores_cleared_grid(tmp_path: Path) -> None:
    class_id, assignment_id = _seed_one_lesson()
    export = client.get("/api/reports/export/all/elementary")
    assert export.status_code == 200, export.text

    with SessionLocal() as session:
        session.execute(delete(ScheduleCell))
        session.commit()

    path = tmp_path / "расписание.xlsx"
    path.write_bytes(export.content)
    with path.open("rb") as fh:
        response = client.post(
            "/api/import/schedule",
            data={"replace": "true"},
            files={
                "file": (
                    "расписание.xlsx",
                    fh,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["placed"] == 1
    assert body["unmatched"] == 0

    with SessionLocal() as session:
        cells = list(session.scalars(select(ScheduleCell)).all())
        assert len(cells) == 1
        cell = cells[0]
        assert cell.class_id == class_id
        assert cell.assignment_id == assignment_id
        assert cell.day_of_week == 1
        assert cell.lesson_number == 1
        assert cell.classroom_id is not None


def test_teacher_export_roundtrip(tmp_path: Path) -> None:
    _seed_one_lesson()
    with SessionLocal() as session:
        teacher_id = session.scalars(select(Teacher)).first().id
    export = client.get(f"/api/reports/export/teacher/{teacher_id}")
    assert export.status_code == 200, export.text

    with SessionLocal() as session:
        session.execute(delete(ScheduleCell))
        session.commit()

    path = tmp_path / "учитель.xlsx"
    path.write_bytes(export.content)
    with path.open("rb") as fh:
        response = client.post(
            "/api/import/schedule",
            files={
                "file": (
                    "учитель.xlsx",
                    fh,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert response.status_code == 200, response.text
    assert response.json()["placed"] == 1

    with SessionLocal() as session:
        cells = list(session.scalars(select(ScheduleCell)).all())
        assert len(cells) == 1
        assert cells[0].lesson_number == 1

    class_id, _ = _seed_one_lesson()
    export = client.get(f"/api/reports/export/class/{class_id}")
    assert export.status_code == 200, export.text

    with SessionLocal() as session:
        session.execute(delete(ScheduleCell))
        session.commit()

    path = tmp_path / "класс.xlsx"
    path.write_bytes(export.content)
    with path.open("rb") as fh:
        response = client.post(
            "/api/import/schedule",
            files={
                "file": (
                    "класс.xlsx",
                    fh,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert response.status_code == 200, response.text
    assert response.json()["placed"] == 1

    with SessionLocal() as session:
        cells = list(session.scalars(select(ScheduleCell)).all())
        assert len(cells) == 1
        assert cells[0].class_id == class_id


def test_import_skips_existing_without_replace(tmp_path: Path) -> None:
    _seed_one_lesson()
    with SessionLocal() as session:
        buf = ReportService(session, TEST_SCHOOL_ID).export_all("elementary").buffer
    path = tmp_path / "grid.xlsx"
    path.write_bytes(buf.getvalue() if hasattr(buf, "getvalue") else buf.read())

    with path.open("rb") as fh:
        response = client.post(
            "/api/import/schedule",
            data={"replace": "false"},
            files={
                "file": (
                    "grid.xlsx",
                    fh,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["placed"] == 0
    assert body["skipped_existing"] == 1

    with SessionLocal() as session:
        assert session.scalars(select(ScheduleCell)).all()
        assert len(list(session.scalars(select(ScheduleCell)).all())) == 1


def test_import_replace_clears_then_restores(tmp_path: Path) -> None:
    class_id, assignment_id = _seed_one_lesson()
    export = client.get("/api/reports/export/all/elementary")
    path = tmp_path / "grid.xlsx"
    path.write_bytes(export.content)

    with SessionLocal() as session:
        cell = session.scalars(select(ScheduleCell)).first()
        assert cell is not None
        cell.lesson_number = 3
        session.commit()

    with path.open("rb") as fh:
        response = client.post(
            "/api/import/schedule",
            data={"replace": "true"},
            files={
                "file": (
                    "grid.xlsx",
                    fh,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["cleared"] == 1
    assert body["placed"] == 1

    with SessionLocal() as session:
        cells = list(session.scalars(select(ScheduleCell)).all())
        assert len(cells) == 1
        assert cells[0].lesson_number == 1
        assert cells[0].assignment_id == assignment_id
        assert cells[0].class_id == class_id


def test_unmatched_subject_is_reported(tmp_path: Path) -> None:
    _seed_one_lesson()
    with SessionLocal() as session:
        session.execute(delete(ScheduleCell))
        session.commit()

    workbook = Workbook()
    ws = workbook.active
    ws.title = "1 смена"
    ws["A1"] = "Понедельник"
    ws["A2"] = "Урок"
    ws["B2"] = "1Б"
    ws["A3"] = "1"
    ws["B3"] = "Неизвестный предмет / каб. 12 / Сидоров С.С."
    path = tmp_path / "bad.xlsx"
    workbook.save(path)

    with path.open("rb") as fh:
        response = client.post(
            "/api/import/schedule",
            files={
                "file": (
                    "bad.xlsx",
                    fh,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["placed"] == 0
    assert body["unmatched"] == 1
    assert any("Неизвестный предмет" in w for w in body["warnings"])


def test_parse_workbook_shift_sheet(tmp_path: Path) -> None:
    path = tmp_path / "shift.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.title = "1 смена"
    ws["A1"] = "Понедельник"
    ws["A2"] = "Урок"
    ws["B2"] = "1Б"
    ws["A3"] = "1\n08:00–08:45"
    ws["B3"] = "Математика / каб. 12 / Сидоров С.С."
    workbook.save(path)
    slots = parse_schedule_workbook(path)
    assert len(slots) == 1
    assert slots[0].class_name == "1Б"
    assert slots[0].day_of_week == 1
    assert slots[0].lesson_number == 1
    assert slots[0].subject == "Математика"


def test_empty_workbook_rejected(tmp_path: Path) -> None:
    path = tmp_path / "empty.xlsx"
    Workbook().save(path)
    with path.open("rb") as fh:
        response = client.post(
            "/api/import/schedule",
            files={
                "file": (
                    "empty.xlsx",
                    fh,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert response.status_code == 400
    assert "нет уроков" in response.text.casefold() or "нет уроков" in str(
        response.json()
    )


def test_subgroups_roundtrip(tmp_path: Path) -> None:
    with SessionLocal() as session:
        shift = Shift(
            school_id=TEST_SCHOOL_ID,
            name="1 смена",
            school_level="elementary",
            start_lesson=1,
            lessons_count=4,
            working_days=5,
            max_lessons_per_day=5,
        )
        subject = Subject(school_id=TEST_SCHOOL_ID, name="Английский язык")
        t1 = Teacher(school_id=TEST_SCHOOL_ID, full_name="Иванова А.А.")
        t2 = Teacher(school_id=TEST_SCHOOL_ID, full_name="Петрова Б.Б.")
        cls = SchoolClass(
            school_id=TEST_SCHOOL_ID, name="1А", grade=1, school_level="elementary"
        )
        r1 = Classroom(school_id=TEST_SCHOOL_ID, number="3")
        r2 = Classroom(school_id=TEST_SCHOOL_ID, number="4")
        session.add_all([shift, subject, t1, t2, cls, r1, r2])
        session.flush()
        cls.shift_id = shift.id
        a1 = TeachingAssignment(
            school_id=TEST_SCHOOL_ID,
            subject_id=subject.id,
            teacher_id=t1.id,
            class_id=cls.id,
            hours_per_week=2,
            group_number=1,
        )
        a2 = TeachingAssignment(
            school_id=TEST_SCHOOL_ID,
            subject_id=subject.id,
            teacher_id=t2.id,
            class_id=cls.id,
            hours_per_week=2,
            group_number=2,
        )
        session.add_all([a1, a2])
        session.flush()
        session.add_all(
            [
                ScheduleCell(
                    school_id=TEST_SCHOOL_ID,
                    class_id=cls.id,
                    day_of_week=2,
                    lesson_number=1,
                    assignment_id=a1.id,
                    classroom_id=r1.id,
                ),
                ScheduleCell(
                    school_id=TEST_SCHOOL_ID,
                    class_id=cls.id,
                    day_of_week=2,
                    lesson_number=1,
                    assignment_id=a2.id,
                    classroom_id=r2.id,
                ),
            ]
        )
        session.commit()

    export = client.get("/api/reports/export/all/elementary")
    assert export.status_code == 200, export.text
    with SessionLocal() as session:
        session.execute(delete(ScheduleCell))
        session.commit()

    path = tmp_path / "groups.xlsx"
    path.write_bytes(export.content)
    with path.open("rb") as fh:
        response = client.post(
            "/api/import/schedule",
            data={"replace": "true"},
            files={
                "file": (
                    "groups.xlsx",
                    fh,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert response.status_code == 200, response.text
    assert response.json()["placed"] == 2

    with SessionLocal() as session:
        cells = list(session.scalars(select(ScheduleCell)).all())
        assert len(cells) == 2
        assert {c.assignment.group_number for c in cells} == {1, 2}
        assert {c.classroom.number for c in cells} == {"3", "4"}
