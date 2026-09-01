"""Excel export of the full timetable includes subject, classroom and teacher."""
from __future__ import annotations

import io

from datetime import time as dt_time

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import delete

from app.models import (
    Classroom,
    Job,
    ScheduleCell,
    ScheduleSettings,
    SchoolClass,
    Shift,
    ShiftLessonTime,
    Subject,
    Teacher,
    TeachingAssignment,
    classroom_subjects,
)
from app.services.report_service import (
    format_grid_export_cell,
    ordinary_lesson_time,
    unique_sheet_name,
)
from backend.deps import SessionLocal
from backend.main import app
from tests.conftest import TEST_SCHOOL_ID

client = TestClient(app)


@pytest.fixture(autouse=True)
def _clear_db() -> None:
    with SessionLocal() as session:
        for model in (
            ScheduleCell,
            TeachingAssignment,
            ShiftLessonTime,
            ScheduleSettings,
            Job,
            SchoolClass,
            Shift,
            classroom_subjects,
            Classroom,
            Subject,
            Teacher,
        ):
            session.execute(delete(model))
        session.commit()


def test_format_grid_export_cell_three_lines() -> None:
    class _Subj:
        display_name = "Математика"

    class _Teacher:
        display_name = "Сидоров С.С."

    class _Room:
        display_name = "12"

    class _Assignment:
        group_number = None

    class _Cell:
        subject = _Subj()
        teacher = _Teacher()
        classroom = _Room()
        assignment = _Assignment()

    text = format_grid_export_cell(_Cell())  # type: ignore[arg-type]
    assert text == "Математика / каб. 12 / Сидоров С.С."


def test_format_grid_export_cell_includes_group() -> None:
    class _Subj:
        display_name = "Математика"

    class _Teacher:
        display_name = "Сидоров С.С."

    class _Room:
        display_name = "12"

    class _Assignment:
        group_number = 2

    class _Cell:
        subject = _Subj()
        teacher = _Teacher()
        classroom = _Room()
        assignment = _Assignment()

    text = format_grid_export_cell(_Cell())  # type: ignore[arg-type]
    assert "Математика · гр.2" in text
    assert "каб. 12" in text
    assert " / " in text


def test_ordinary_lesson_time_skips_class_hour_day() -> None:
    times = {
        1: {1: "08:25–09:10"},
        2: {1: "08:00–08:45"},
    }
    assert ordinary_lesson_time(times, 1, 5, 1) == "08:00–08:45"
    assert ordinary_lesson_time(times, 1, 5, None) == "08:25–09:10"


def test_export_all_cell_contains_subject_classroom_teacher() -> None:
    with SessionLocal() as session:
        shift = Shift(
            school_id=TEST_SCHOOL_ID,
            name="1 смена",
            school_level="elementary",
            start_lesson=1,
            lessons_count=5,
            working_days=5,
            max_lessons_per_day=5,
        )
        subject = Subject(school_id=TEST_SCHOOL_ID, name="Математика")
        teacher = Teacher(school_id=TEST_SCHOOL_ID, full_name="Сидоров С.С.")
        cls = SchoolClass(
            school_id=TEST_SCHOOL_ID, name="1Б", grade=1, school_level="elementary"
        )
        room = Classroom(school_id=TEST_SCHOOL_ID, number="12")
        session.add_all([shift, subject, teacher, cls, room])
        session.flush()
        cls.shift_id = shift.id
        assignment = TeachingAssignment(
            school_id=TEST_SCHOOL_ID,
            subject_id=subject.id,
            teacher_id=teacher.id,
            class_id=cls.id,
            hours_per_week=4,
        )
        session.add(assignment)
        session.flush()
        session.add(
            ScheduleCell(
                school_id=TEST_SCHOOL_ID,
                class_id=cls.id,
                day_of_week=1,
                lesson_number=1,
                assignment_id=assignment.id,
                classroom_id=room.id,
            )
        )
        session.commit()

    response = client.get("/api/reports/export/all/elementary")
    assert response.status_code == 200, response.text
    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["1 смена"]
    sheet = workbook["1 смена"]
    # Row 1 = day title, row 2 = headers, row 3 = lesson 1
    assert sheet["A1"].value == "Понедельник"
    assert sheet["B2"].value == "1Б"
    cell_value = sheet["B3"].value
    assert cell_value is not None
    assert "Математика" in cell_value
    assert "каб. 12" in cell_value
    assert "Сидоров С.С." in cell_value
    assert " / " in cell_value
    assert sheet["B3"].alignment.wrap_text is True
    assert sheet["B3"].alignment.horizontal == "left"
    assert sheet["B3"].font.size == 8
    assert sheet["B3"].border.left.style == "medium"
    assert sheet["B3"].border.bottom.style == "medium"
    assert sheet.row_dimensions[3].height >= 24
    assert sheet["A1"].fill.patternType == "solid"
    assert sheet["B2"].fill.patternType == "solid"
    assert sheet["B3"].fill.patternType == "solid"
    assert sheet["A3"].fill.patternType == "solid"


def test_class_report_and_export_show_subject_room_teacher() -> None:
    with SessionLocal() as session:
        shift = Shift(
            school_id=TEST_SCHOOL_ID,
            name="1 смена",
            school_level="elementary",
            start_lesson=1,
            lessons_count=5,
            working_days=5,
            max_lessons_per_day=5,
        )
        subject = Subject(school_id=TEST_SCHOOL_ID, name="Математика")
        teacher = Teacher(school_id=TEST_SCHOOL_ID, full_name="Сидоров С.С.")
        cls = SchoolClass(
            school_id=TEST_SCHOOL_ID, name="1Б", grade=1, school_level="elementary"
        )
        room = Classroom(school_id=TEST_SCHOOL_ID, number="12")
        session.add_all([shift, subject, teacher, cls, room])
        session.flush()
        cls.shift_id = shift.id
        assignment = TeachingAssignment(
            school_id=TEST_SCHOOL_ID,
            subject_id=subject.id,
            teacher_id=teacher.id,
            class_id=cls.id,
            hours_per_week=4,
        )
        session.add(assignment)
        session.flush()
        session.add(
            ScheduleCell(
                school_id=TEST_SCHOOL_ID,
                class_id=cls.id,
                day_of_week=1,
                lesson_number=1,
                assignment_id=assignment.id,
                classroom_id=room.id,
            )
        )
        class_id = cls.id
        session.commit()

    report = client.get(f"/api/reports/class/{class_id}")
    assert report.status_code == 200, report.text
    cells = report.json()["cells"]
    assert len(cells) == 1
    assert cells[0]["subject_name"] == "Математика"
    assert cells[0]["teacher_name"] == "Сидоров С.С."
    assert cells[0]["classroom_name"] == "12"

    xlsx = client.get(f"/api/reports/export/class/{class_id}")
    assert xlsx.status_code == 200, xlsx.text
    sheet = load_workbook(io.BytesIO(xlsx.content)).active
    assert sheet["A1"].value == "Урок"
    assert sheet["B1"].value == "Понедельник"
    monday = sheet["B2"].value
    assert monday is not None
    assert "Математика" in monday
    assert "каб. 12" in monday
    assert "Сидоров С.С." in monday
    assert " / " in monday
    assert sheet["B2"].alignment.wrap_text is True
    assert sheet["B2"].alignment.horizontal == "left"
    assert sheet["B2"].font.size == 8
    assert sheet["B2"].border.left.style == "medium"
    assert sheet["B2"].fill.patternType == "solid"
    assert sheet["A1"].fill.patternType == "solid"


def test_class_and_teacher_reports_include_class_hour_and_bells() -> None:
    with SessionLocal() as session:
        shift = Shift(
            school_id=TEST_SCHOOL_ID,
            name="1 смена",
            school_level="elementary",
            start_lesson=1,
            lessons_count=5,
            working_days=5,
            max_lessons_per_day=5,
            class_hour_day=1,
            class_hour_start=dt_time(8, 0),
            class_hour_end=dt_time(8, 20),
        )
        subject = Subject(school_id=TEST_SCHOOL_ID, name="Математика")
        teacher = Teacher(school_id=TEST_SCHOOL_ID, full_name="Сидоров С.С.")
        cls = SchoolClass(
            school_id=TEST_SCHOOL_ID, name="1Б", grade=1, school_level="elementary"
        )
        session.add_all([shift, subject, teacher, cls])
        session.flush()
        cls.shift_id = shift.id
        session.add_all(
            [
                ShiftLessonTime(
                    school_id=TEST_SCHOOL_ID,
                    shift_id=shift.id,
                    day_of_week=1,
                    lesson_number=1,
                    time_start=dt_time(8, 25),
                    time_end=dt_time(9, 10),
                ),
                ShiftLessonTime(
                    school_id=TEST_SCHOOL_ID,
                    shift_id=shift.id,
                    day_of_week=2,
                    lesson_number=1,
                    time_start=dt_time(8, 0),
                    time_end=dt_time(8, 45),
                ),
            ]
        )
        assignment = TeachingAssignment(
            school_id=TEST_SCHOOL_ID,
            subject_id=subject.id,
            teacher_id=teacher.id,
            class_id=cls.id,
            hours_per_week=4,
        )
        session.add(assignment)
        session.flush()
        session.add(
            ScheduleCell(
                school_id=TEST_SCHOOL_ID,
                class_id=cls.id,
                day_of_week=1,
                lesson_number=1,
                assignment_id=assignment.id,
            )
        )
        class_id, teacher_id = cls.id, teacher.id
        session.commit()

    class_body = client.get(f"/api/reports/class/{class_id}").json()
    assert class_body["class_hour_day"] == 1
    assert class_body["class_hour_time_label"] == "08:00–08:20"
    assert class_body["lesson_times_by_day"]["1"]["1"] == "08:25–09:10"
    assert class_body["lesson_times_by_day"]["2"]["1"] == "08:00–08:45"

    teacher_body = client.get(f"/api/reports/teacher/{teacher_id}").json()
    assert teacher_body["class_hour_day"] == 1
    assert teacher_body["class_hour_time_label"] == "08:00–08:20"
    assert teacher_body["lesson_times_by_day"]["1"]["1"] == "08:25–09:10"
    assert teacher_body["lesson_times_by_day"]["2"]["1"] == "08:00–08:45"
    assert teacher_body["lessons_range"] == [1, 2, 3, 4, 5]

    xlsx = client.get(f"/api/reports/export/class/{class_id}")
    assert xlsx.status_code == 200, xlsx.text
    sheet = load_workbook(io.BytesIO(xlsx.content)).active
    # Row 1 headers, row 2 class hour, row 3 lesson 1
    lesson_col = str(sheet["A3"].value or "")
    assert "1" in lesson_col
    assert "08:00–08:45" in lesson_col
    assert "08:25–09:10" not in lesson_col
    monday = str(sheet["B3"].value or "")
    assert "08:25–09:10" in monday
    assert "Математика" in monday
    tuesday = str(sheet["C3"].value or "")
    assert "08:00–08:45" not in tuesday


def test_unique_sheet_name_sanitizes_and_dedupes() -> None:
    used: set[str] = set()
    assert unique_sheet_name("1 смена", used) == "1 смена"
    assert unique_sheet_name("1 смена", used) == "1 смена (2)"
    assert unique_sheet_name("Пн/Вт", used) == "Пн Вт"


def test_export_all_splits_sheets_by_shift() -> None:
    with SessionLocal() as session:
        shift1 = Shift(
            school_id=TEST_SCHOOL_ID,
            name="1 смена",
            school_level="elementary",
            start_lesson=1,
            lessons_count=4,
            working_days=5,
            max_lessons_per_day=5,
        )
        shift2 = Shift(
            school_id=TEST_SCHOOL_ID,
            name="2 смена",
            school_level="elementary",
            start_lesson=1,
            lessons_count=4,
            working_days=5,
            max_lessons_per_day=5,
        )
        subject = Subject(school_id=TEST_SCHOOL_ID, name="Русский язык")
        t1 = Teacher(school_id=TEST_SCHOOL_ID, full_name="Иванова А.А.")
        t2 = Teacher(school_id=TEST_SCHOOL_ID, full_name="Петрова Б.Б.")
        c1 = SchoolClass(
            school_id=TEST_SCHOOL_ID, name="1А", grade=1, school_level="elementary"
        )
        c2 = SchoolClass(
            school_id=TEST_SCHOOL_ID, name="1Б", grade=1, school_level="elementary"
        )
        room1 = Classroom(school_id=TEST_SCHOOL_ID, number="10")
        room2 = Classroom(school_id=TEST_SCHOOL_ID, number="20")
        session.add_all([shift1, shift2, subject, t1, t2, c1, c2, room1, room2])
        session.flush()
        c1.shift_id = shift1.id
        c2.shift_id = shift2.id
        a1 = TeachingAssignment(
            school_id=TEST_SCHOOL_ID,
            subject_id=subject.id,
            teacher_id=t1.id,
            class_id=c1.id,
            hours_per_week=3,
        )
        a2 = TeachingAssignment(
            school_id=TEST_SCHOOL_ID,
            subject_id=subject.id,
            teacher_id=t2.id,
            class_id=c2.id,
            hours_per_week=3,
        )
        session.add_all([a1, a2])
        session.flush()
        session.add_all(
            [
                ScheduleCell(
                    school_id=TEST_SCHOOL_ID,
                    class_id=c1.id,
                    day_of_week=1,
                    lesson_number=1,
                    assignment_id=a1.id,
                    classroom_id=room1.id,
                ),
                ScheduleCell(
                    school_id=TEST_SCHOOL_ID,
                    class_id=c2.id,
                    day_of_week=2,
                    lesson_number=1,
                    assignment_id=a2.id,
                    classroom_id=room2.id,
                ),
            ]
        )
        session.commit()

    response = client.get("/api/reports/export/all/elementary")
    assert response.status_code == 200, response.text
    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["1 смена", "2 смена"]

    first = workbook["1 смена"]
    assert first["B2"].value == "1А"
    assert "1Б" not in {first.cell(2, col).value for col in range(1, 6)}
    monday = first["B3"].value
    assert monday is not None
    assert "Русский язык" in monday
    assert "каб. 10" in monday
    assert "Иванова А.А." in monday

    second = workbook["2 смена"]
    assert second["B2"].value == "1Б"
    assert "1А" not in {second.cell(2, col).value for col in range(1, 6)}
    # Tuesday block: 1 title + 1 header + 4 lessons + 1 blank = 7 rows per day
    tuesday_header_row = 1 + (1 + 1 + 4 + 1)
    assert second.cell(tuesday_header_row, 1).value == "Вторник"
    tuesday_lesson = second.cell(tuesday_header_row + 2, 2).value
    assert tuesday_lesson is not None
    assert "Русский язык" in tuesday_lesson
    assert "каб. 20" in tuesday_lesson
    assert "Петрова Б.Б." in tuesday_lesson
