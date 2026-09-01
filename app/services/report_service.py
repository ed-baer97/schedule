"""Reports and Excel export business logic."""
from __future__ import annotations

import io
from collections import defaultdict
from dataclasses import dataclass

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.domain import DAY_NAMES, lesson_end_exclusive, time_range_label
from app.models import ScheduleCell, SchoolClass, Shift, Subject, Teacher, TeachingAssignment
from app.services.errors import BadRequestError
from app.services.schedule_mapping import (
    CELL_LOAD_WITH_CLASS,
    cell_to_report_dict,
    load_cells,
)
from app.services.tenancy import require_owned

_INK = "14201A"
_MUTED = "3F5248"
_WHITE = "F4FFFD"
_TEAL = "147F78"
_TEAL_DEEP = "0E5C57"
_LINE = "2C3A34"
_EXCEL_SHEET_BAD = frozenset("\\/*?:[]")
_CELL_FONT_SIZE = 8
_LINE_PT = 12

_GRID = Border(
    left=Side(style="medium", color=_LINE),
    right=Side(style="medium", color=_LINE),
    top=Side(style="medium", color=_LINE),
    bottom=Side(style="medium", color=_LINE),
)
_TITLE_FILL = PatternFill("solid", fgColor=_TEAL_DEEP)
_HEADER_FILL = PatternFill("solid", fgColor=_TEAL)
_LESSON_COL_FILL = PatternFill("solid", fgColor="DCE8E4")
_EMPTY_FILL = PatternFill("solid", fgColor="F7FAF8")
_CLASS_HOUR_FILL = PatternFill("solid", fgColor="F6E8D4")

_TITLE_FONT = Font(bold=True, size=11, color=_WHITE, name="Calibri")
_HEADER_FONT = Font(bold=True, size=9, color=_WHITE, name="Calibri")
_LESSON_NUM_FONT = Font(bold=True, size=_CELL_FONT_SIZE, color=_INK, name="Calibri")
_SUBJECT_FONT = Font(size=_CELL_FONT_SIZE, color=_INK, name="Calibri")
_META_FONT = Font(size=_CELL_FONT_SIZE, color=_MUTED, name="Calibri")
_CLASS_HOUR_FONT = Font(size=_CELL_FONT_SIZE, color=_INK, name="Calibri")

_TITLE_ALIGN = Alignment(horizontal="left", vertical="center", indent=1)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_NUM_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
_CELL_ALIGN = Alignment(
    wrap_text=True,
    shrink_to_fit=True,
    vertical="center",
    horizontal="left",
)


def format_grid_export_cell(cell: ScheduleCell) -> str:
    """Subject, classroom and teacher — compact slash-separated cell text."""
    subject = cell.subject.display_name if cell.subject else "?"
    if cell.assignment and cell.assignment.group_number:
        subject = f"{subject} · гр.{cell.assignment.group_number}"
    room = cell.classroom.display_name if cell.classroom else "—"
    teacher = cell.teacher.display_name if cell.teacher else "—"
    return f"{subject} / каб. {room} / {teacher}"


def _lighten_hex(color: str, white: float = 0.78) -> str:
    raw = (color or Subject.DEFAULT_COLOR).lstrip("#")
    if len(raw) != 6:
        raw = Subject.DEFAULT_COLOR.lstrip("#")
    try:
        r, g, b = int(raw[0:2], 16), int(raw[2:4], 16), int(raw[4:6], 16)
    except ValueError:
        r, g, b = 20, 127, 120
    r = round(r * (1 - white) + 255 * white)
    g = round(g * (1 - white) + 255 * white)
    b = round(b * (1 - white) + 255 * white)
    return f"{r:02X}{g:02X}{b:02X}"


def _subject_fill(cells: list[ScheduleCell]) -> PatternFill:
    hex_color = Subject.DEFAULT_COLOR
    if cells:
        subj = cells[0].subject
        if subj is not None:
            hex_color = subj.display_color or Subject.DEFAULT_COLOR
    return PatternFill("solid", fgColor=_lighten_hex(hex_color))


def _paint(cell, *, fill, font, alignment, border=_GRID) -> None:
    cell.fill = fill
    cell.font = font
    cell.alignment = alignment
    cell.border = border


def _write_grid_value(
    ws: Worksheet,
    row: int,
    col: int,
    match: list[ScheduleCell],
    time_label: str | None = None,
) -> None:
    parts: list[str] = []
    if time_label:
        parts.append(time_label)
    if match:
        parts.extend(format_grid_export_cell(c) for c in match)
    value = "\n".join(parts)
    excel_cell = ws.cell(row, col, value)
    if match:
        _paint(
            excel_cell,
            fill=_subject_fill(match),
            font=_SUBJECT_FONT,
            alignment=_CELL_ALIGN,
        )
    else:
        _paint(excel_cell, fill=_EMPTY_FILL, font=_META_FONT, alignment=_CELL_ALIGN)


def ordinary_lesson_time(
    times_by_day: dict[int, dict[int, str]],
    lesson: int,
    working_days: int,
    class_hour_day: int | None,
) -> str | None:
    """Bell label for days without a class hour; falls back to class-hour day."""
    ordered = list(range(1, working_days + 1))
    if class_hour_day:
        ordered = [d for d in ordered if d != class_hour_day] + [class_hour_day]
    for day in ordered:
        label = times_by_day.get(day, {}).get(lesson)
        if label:
            return label
    return None


def _lesson_index_label(lesson: int, time_label: str | None) -> str:
    return f"{lesson}\n{time_label}" if time_label else str(lesson)


def _lesson_row_height(max_groups: int) -> float:
    """Room for one slash-line per group, plus a wrap line if names are long."""
    lines = max(2, max_groups * 2)
    return 6 + _LINE_PT * lines


def _style_sheet_chrome(ws: Worksheet, last_col: int, lesson_col_width: float = 12) -> None:
    ws.column_dimensions["A"].width = lesson_col_width
    for col_idx in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_options.horizontalCentered = True


def unique_sheet_name(raw: str, used: set[str]) -> str:
    cleaned = "".join(" " if ch in _EXCEL_SHEET_BAD else ch for ch in (raw or "").strip())
    cleaned = " ".join(cleaned.split()) or "Смена"
    base = cleaned[:31]
    name = base
    n = 2
    while name.casefold() in {u.casefold() for u in used}:
        suffix = f" ({n})"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


def _shift_lesson_numbers(shift: Shift | None) -> list[int]:
    if shift:
        start = shift.start_lesson or 1
        count = shift.lessons_count or 6
        return list(range(start, start + count))
    return list(range(1, 8))


def _write_shift_sheet(
    ws: Worksheet,
    classes: list[SchoolClass],
    shift: Shift | None,
    cell_index: dict[tuple[int, int, int], list[ScheduleCell]],
) -> None:
    bells = shift_bell_fields(shift)
    times_by_day: dict[int, dict[int, str]] = bells["lesson_times_by_day"]
    working_days = shift.working_days if shift else 5
    lessons = _shift_lesson_numbers(shift)
    class_count = len(classes)
    last_col = max(1, 1 + class_count)
    row = 1

    for day in range(1, working_days + 1):
        for col in range(1, last_col + 1):
            title_cell = ws.cell(row, col, DAY_NAMES[day - 1] if col == 1 else None)
            _paint(
                title_cell,
                fill=_TITLE_FILL,
                font=_TITLE_FONT,
                alignment=_TITLE_ALIGN,
            )
        if last_col > 1:
            ws.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=last_col
            )
        ws.row_dimensions[row].height = 18
        row += 1

        header = ws.cell(row, 1, "Урок")
        _paint(header, fill=_HEADER_FILL, font=_HEADER_FONT, alignment=_HEADER_ALIGN)
        for i, school_class in enumerate(classes):
            class_header = ws.cell(row, 2 + i, school_class.name)
            _paint(
                class_header,
                fill=_HEADER_FILL,
                font=_HEADER_FONT,
                alignment=_HEADER_ALIGN,
            )
        ws.row_dimensions[row].height = 16
        row += 1

        if bells["class_hour_day"] == day:
            ch_time = bells["class_hour_time_label"]
            ch_label = f"Кл. час\n{ch_time}" if ch_time else "Кл. час"
            index_cell = ws.cell(row, 1, ch_label)
            _paint(
                index_cell,
                fill=_CLASS_HOUR_FILL,
                font=_LESSON_NUM_FONT,
                alignment=_NUM_ALIGN,
            )
            for i in range(class_count):
                hour_cell = ws.cell(row, 2 + i, "Классный час")
                _paint(
                    hour_cell,
                    fill=_CLASS_HOUR_FILL,
                    font=_CLASS_HOUR_FONT,
                    alignment=_CELL_ALIGN,
                )
            ws.row_dimensions[row].height = 24
            row += 1

        day_times = times_by_day.get(day, {})
        day_lessons = (
            list(range(shift.start_lesson, lesson_end_exclusive(shift, day)))
            if shift
            else lessons
        )
        for lesson in day_lessons:
            index_cell = ws.cell(
                row, 1, _lesson_index_label(lesson, day_times.get(lesson))
            )
            _paint(
                index_cell,
                fill=_LESSON_COL_FILL,
                font=_LESSON_NUM_FONT,
                alignment=_NUM_ALIGN,
            )
            max_groups = 1
            for i, school_class in enumerate(classes):
                match = cell_index.get((school_class.id, day, lesson), [])
                max_groups = max(max_groups, len(match) or 1)
                _write_grid_value(ws, row, 2 + i, match)
            ws.row_dimensions[row].height = _lesson_row_height(max_groups)
            row += 1

        row += 1

    _style_sheet_chrome(ws, last_col)


def shift_bell_fields(shift: Shift | None) -> dict:
    """Lesson bells and class-hour label for a shift (empty if none)."""
    lesson_times_by_day: dict[int, dict[int, str]] = {}
    class_hour_time_label: str | None = None
    class_hour_day: int | None = None
    if shift:
        class_hour_day = shift.class_hour_day
        class_hour_time_label = time_range_label(
            shift.class_hour_start, shift.class_hour_end
        )
        for lt in shift.lesson_times.all():
            label = time_range_label(lt.time_start, lt.time_end)
            if label:
                lesson_times_by_day.setdefault(lt.day_of_week, {})[lt.lesson_number] = (
                    label
                )
    return {
        "lesson_times_by_day": lesson_times_by_day,
        "class_hour_day": class_hour_day,
        "class_hour_time_label": class_hour_time_label,
    }


@dataclass(frozen=True)
class ExportFile:
    buffer: io.BytesIO
    filename: str


class ReportService:
    def __init__(self, db: Session, school_id: int):
        self.db = db
        self.school_id = school_id

    def class_report(self, class_id: int) -> dict:
        school_class = require_owned(self.db, SchoolClass, class_id, self.school_id)

        shift = school_class.shift if school_class.shift_id else None
        working_days = shift.working_days if shift else 5
        max_lessons = shift.max_lessons_per_day if shift else 7
        lessons_range = (
            list(range(shift.start_lesson, shift.start_lesson + shift.lessons_count))
            if shift
            else list(range(1, max_lessons + 1))
        )
        bells = shift_bell_fields(shift)

        cells = load_cells(
            self.db,
            ScheduleCell.class_id == class_id,
            ScheduleCell.school_id == self.school_id,
            with_class=True,
        )
        return {
            "class_id": school_class.id,
            "class_name": school_class.name,
            "school_level": school_class.school_level,
            "day_names": DAY_NAMES,
            "working_days": working_days,
            "max_lessons": max_lessons,
            "lessons_range": lessons_range,
            "class_hour_day": bells["class_hour_day"],
            "class_hour_time_label": bells["class_hour_time_label"],
            "lesson_times_by_day": bells["lesson_times_by_day"],
            "cells": [cell_to_report_dict(c) for c in cells],
        }

    def teacher_report(self, teacher_id: int) -> dict:
        teacher = require_owned(self.db, Teacher, teacher_id, self.school_id)

        cells = list(
            self.db.execute(
                select(ScheduleCell)
                .join(TeachingAssignment)
                .options(*CELL_LOAD_WITH_CLASS)
                .where(
                    TeachingAssignment.teacher_id == teacher_id,
                    ScheduleCell.school_id == self.school_id,
                )
            )
            .scalars()
            .unique()
            .all()
        )
        working_days = 5
        max_lessons = 7
        shift_counts: dict[int, int] = {}
        shift_by_id: dict[int, Shift] = {}
        for cell in cells:
            sc = cell.school_class
            sh = sc.shift if sc is not None and sc.shift_id else None
            if sh:
                working_days = max(working_days, sh.working_days)
                max_lessons = max(max_lessons, sh.max_lessons_per_day)
                shift_counts[sh.id] = shift_counts.get(sh.id, 0) + 1
                shift_by_id[sh.id] = sh

        shift = None
        if shift_counts:
            best_id = max(shift_counts, key=lambda sid: shift_counts[sid])
            shift = shift_by_id[best_id]
        bells = shift_bell_fields(shift)
        lessons_range = (
            list(range(shift.start_lesson, shift.start_lesson + shift.lessons_count))
            if shift
            else list(range(1, max_lessons + 1))
        )

        return {
            "teacher_id": teacher.id,
            "teacher_name": teacher.full_name,
            "day_names": DAY_NAMES,
            "working_days": working_days,
            "max_lessons": max_lessons,
            "lessons_range": lessons_range,
            "class_hour_day": bells["class_hour_day"],
            "class_hour_time_label": bells["class_hour_time_label"],
            "lesson_times_by_day": bells["lesson_times_by_day"],
            "cells": [cell_to_report_dict(c) for c in cells],
        }

    def export_class(self, class_id: int) -> ExportFile:
        school_class = require_owned(self.db, SchoolClass, class_id, self.school_id)
        shift = school_class.shift if school_class.shift_id else None
        working_days = shift.working_days if shift else 5
        lessons = _shift_lesson_numbers(shift)

        times_by_day: dict[int, dict[int, str]] = {}
        if shift:
            for lt in shift.lesson_times.all():
                label = time_range_label(lt.time_start, lt.time_end)
                if label:
                    times_by_day.setdefault(lt.day_of_week, {})[lt.lesson_number] = label

        cells = load_cells(
            self.db,
            ScheduleCell.class_id == class_id,
            ScheduleCell.school_id == self.school_id,
            with_class=True,
        )
        cell_index: dict[tuple[int, int], list[ScheduleCell]] = defaultdict(list)
        for cell in cells:
            cell_index[(cell.day_of_week, cell.lesson_number)].append(cell)

        workbook = Workbook()
        ws = workbook.active
        ws.title = unique_sheet_name(f"Расписание {school_class.name}", set())
        last_col = 1 + working_days
        bells = shift_bell_fields(shift)

        header = ws.cell(1, 1, "Урок")
        _paint(header, fill=_HEADER_FILL, font=_HEADER_FONT, alignment=_HEADER_ALIGN)
        for day in range(1, working_days + 1):
            day_header = ws.cell(1, 1 + day, DAY_NAMES[day - 1])
            _paint(
                day_header,
                fill=_HEADER_FILL,
                font=_HEADER_FONT,
                alignment=_HEADER_ALIGN,
            )
        ws.row_dimensions[1].height = 16
        row_idx = 2

        if bells["class_hour_day"]:
            ch_time = bells["class_hour_time_label"]
            ch_label = f"Кл. час\n{ch_time}" if ch_time else "Кл. час"
            index_cell = ws.cell(row_idx, 1, ch_label)
            _paint(
                index_cell,
                fill=_CLASS_HOUR_FILL,
                font=_LESSON_NUM_FONT,
                alignment=_NUM_ALIGN,
            )
            for day in range(1, working_days + 1):
                hour_text = (
                    "Классный час" if bells["class_hour_day"] == day else ""
                )
                hour_cell = ws.cell(row_idx, 1 + day, hour_text)
                _paint(
                    hour_cell,
                    fill=_CLASS_HOUR_FILL if hour_text else _EMPTY_FILL,
                    font=_CLASS_HOUR_FONT,
                    alignment=_CELL_ALIGN,
                )
            ws.row_dimensions[row_idx].height = 24
            row_idx += 1

        for lesson in lessons:
            time_label = ordinary_lesson_time(
                times_by_day,
                lesson,
                working_days,
                bells["class_hour_day"],
            )
            index_cell = ws.cell(row_idx, 1, _lesson_index_label(lesson, time_label))
            _paint(
                index_cell,
                fill=_LESSON_COL_FILL,
                font=_LESSON_NUM_FONT,
                alignment=_NUM_ALIGN,
            )
            max_groups = 1
            extra_time = False
            for day in range(1, working_days + 1):
                match = cell_index.get((day, lesson), [])
                max_groups = max(max_groups, len(match) or 1)
                day_time = times_by_day.get(day, {}).get(lesson)
                day_bell = day_time if day_time and day_time != time_label else None
                if day_bell:
                    extra_time = True
                _write_grid_value(ws, row_idx, 1 + day, match, day_bell)
            ws.row_dimensions[row_idx].height = _lesson_row_height(
                max_groups + (1 if extra_time else 0)
            )
            row_idx += 1

        _style_sheet_chrome(ws, last_col)

        buf = io.BytesIO()
        workbook.save(buf)
        return ExportFile(buf, f"расписание_{school_class.name}.xlsx")

    def export_teacher(self, teacher_id: int) -> ExportFile:
        teacher = require_owned(self.db, Teacher, teacher_id, self.school_id)
        cells = list(
            self.db.execute(
                select(ScheduleCell)
                .join(TeachingAssignment)
                .options(*CELL_LOAD_WITH_CLASS)
                .where(
                    TeachingAssignment.teacher_id == teacher_id,
                    ScheduleCell.school_id == self.school_id,
                )
            )
            .scalars()
            .unique()
            .all()
        )
        working_days = 5
        max_lessons = 7
        for cell in cells:
            sh = cell.school_class.shift if cell.school_class.shift_id else None
            if sh:
                working_days = max(working_days, sh.working_days)
                max_lessons = max(max_lessons, sh.max_lessons_per_day)

        data = []
        for day in range(1, working_days + 1):
            for lesson in range(1, max_lessons + 1):
                row = {"День": DAY_NAMES[day - 1], "Урок": lesson}
                match = [
                    c for c in cells if c.day_of_week == day and c.lesson_number == lesson
                ]
                if match:
                    row["Класс"] = " / ".join(c.school_class.name for c in match)
                    row["Предмет"] = " / ".join(c.subject.display_name for c in match)
                    row["Кабинет"] = " / ".join(
                        (c.classroom.number if c.classroom else "—") for c in match
                    )
                else:
                    row["Класс"] = ""
                    row["Предмет"] = ""
                    row["Кабинет"] = ""
                data.append(row)

        df = pd.DataFrame(data)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(
                writer, sheet_name=f"Расписание {teacher.full_name}"[:31], index=False
            )
        filename = teacher.full_name.replace(" ", "_").replace(".", "")
        return ExportFile(buf, f"расписание_{filename}.xlsx")

    def export_all(self, school_level: str) -> ExportFile:
        if school_level not in ("elementary", "secondary"):
            raise BadRequestError("Invalid school_level")
        classes = list(
            self.db.scalars(
                select(SchoolClass)
                .where(
                    SchoolClass.school_level == school_level,
                    SchoolClass.school_id == self.school_id,
                )
                .order_by(SchoolClass.grade, SchoolClass.name)
            ).all()
        )
        shifts = list(
            self.db.scalars(
                select(Shift)
                .where(
                    Shift.school_id == self.school_id,
                    Shift.school_level == school_level,
                )
                .order_by(Shift.start_lesson, Shift.name)
            ).all()
        )

        by_shift: dict[int, list[SchoolClass]] = {s.id: [] for s in shifts}
        unassigned: list[SchoolClass] = []
        for school_class in classes:
            if school_class.shift_id in by_shift:
                by_shift[school_class.shift_id].append(school_class)
            else:
                unassigned.append(school_class)

        groups: list[tuple[str, list[SchoolClass], Shift | None]] = [
            (shift.name, by_shift[shift.id], shift) for shift in shifts
        ]
        if unassigned:
            groups.append(("Без смены", unassigned, None))
        if not groups:
            groups.append(("Расписание", [], None))

        class_ids = [c.id for c in classes]
        all_cells = (
            load_cells(
                self.db,
                ScheduleCell.class_id.in_(class_ids),
                ScheduleCell.school_id == self.school_id,
                with_class=True,
            )
            if class_ids
            else []
        )
        cell_index: dict[tuple[int, int, int], list[ScheduleCell]] = defaultdict(list)
        for cell in all_cells:
            cell_index[(cell.class_id, cell.day_of_week, cell.lesson_number)].append(
                cell
            )

        workbook = Workbook()
        used_names: set[str] = set()
        for index, (raw_name, shift_classes, shift) in enumerate(groups):
            sheet_name = unique_sheet_name(raw_name, used_names)
            if index == 0:
                ws = workbook.active
                ws.title = sheet_name
            else:
                ws = workbook.create_sheet(sheet_name)
            _write_shift_sheet(ws, shift_classes, shift, cell_index)

        buf = io.BytesIO()
        workbook.save(buf)
        level_name = "начальная" if school_level == "elementary" else "основная"
        return ExportFile(buf, f"расписание_{level_name}_школа.xlsx")
