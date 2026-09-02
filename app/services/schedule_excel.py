"""Restore schedule cells from a timetable Excel export."""
from __future__ import annotations

import re
from dataclasses import dataclass, field

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.domain import DAY_NAMES, normalize_person_name
from app.models import Classroom, ScheduleCell, SchoolClass, Subject, Teacher, TeachingAssignment
from app.services.schedule.service import ScheduleService
from app.services.schedule.types import Placement

_CLASS_NAME_RE = re.compile(r"^(\d{1,2})\s*([^\W\d_]{1,3})$", re.UNICODE)
_TIME_LINE_RE = re.compile(r"^\d{1,2}:\d{2}\s*[–—-]\s*\d{1,2}:\d{2}$")
_GROUP_SUFFIX_RE = re.compile(r"\s*[·(]\s*гр\.?\s*(\d+)\s*\)?\s*$", re.IGNORECASE)
_ROOM_PREFIX_RE = re.compile(r"^каб\.?\s*", re.IGNORECASE)
_EMPTY_MARKS = frozenset({"", "—", "-", "–", "?", "nan"})
_DAY_BY_NAME = {name.casefold(): index + 1 for index, name in enumerate(DAY_NAMES)}
_SKIP_INDEX_PREFIXES = ("перемена", "кл. час", "кл.час", "классный час")


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm_class(name: str) -> str:
    return name.replace(" ", "").casefold()


def _is_class_name(text: str) -> bool:
    return bool(text and _CLASS_NAME_RE.match(text.replace(" ", "")))


def _day_number(text: str) -> int | None:
    return _DAY_BY_NAME.get((text or "").casefold())


def _lesson_number(index_text: str) -> int | None:
    first = (index_text or "").split("\n", 1)[0].strip()
    if not first.isdigit():
        return None
    return int(first)


def _is_skip_index(text: str) -> bool:
    folded = (text or "").split("\n", 1)[0].strip().casefold()
    return any(folded.startswith(prefix) for prefix in _SKIP_INDEX_PREFIXES)


def _empty_to_none(text: str | None) -> str | None:
    value = (text or "").strip()
    if value.casefold() in _EMPTY_MARKS:
        return None
    return value


def _room_name(text: str) -> str | None:
    return _empty_to_none(_ROOM_PREFIX_RE.sub("", text).strip())


def _split_group(text: str) -> tuple[str, int | None]:
    raw = (text or "").strip()
    match = _GROUP_SUFFIX_RE.search(raw)
    if not match:
        return raw, None
    return raw[: match.start()].strip(), int(match.group(1))


@dataclass(frozen=True)
class CellEntry:
    subject: str
    teacher: str | None = None
    classroom: str | None = None
    group_number: int | None = None
    class_name: str | None = None


@dataclass(frozen=True)
class ParsedSlot:
    class_name: str
    day_of_week: int
    lesson_number: int
    subject: str
    teacher: str | None = None
    classroom: str | None = None
    group_number: int | None = None


@dataclass
class ImportScheduleResultData:
    placed: int
    skipped_existing: int
    unmatched: int
    cleared: int
    warnings: list[str] = field(default_factory=list)
    message: str = ""


def parse_cell_entries(text: str) -> list[CellEntry]:
    """Parse one exported timetable cell into lesson entries."""
    if not text:
        return []
    lines = [line.strip() for line in str(text).splitlines() if line.strip()]
    if lines and _TIME_LINE_RE.match(lines[0]):
        lines = lines[1:]
    lines = [line for line in lines if not _TIME_LINE_RE.match(line)]
    if not lines:
        return []
    if any(line.casefold() in {"классный час"} for line in lines) and len(lines) == 1:
        return []
    if any(" / " in line for line in lines):
        entries: list[CellEntry] = []
        for line in lines:
            parsed = _parse_slash_line(line)
            if parsed is not None:
                entries.append(parsed)
        return entries
    return _parse_block_lines(lines)


def _parse_slash_line(line: str) -> CellEntry | None:
    parts = [part.strip() for part in line.split(" / ")]
    if len(parts) < 2:
        return None
    last_is_room = bool(_ROOM_PREFIX_RE.match(parts[-1]))
    first_is_class = _is_class_name(parts[0])
    if last_is_room and first_is_class and len(parts) >= 3:
        subject, group = _split_group(parts[1])
        if not subject:
            return None
        return CellEntry(
            subject=subject,
            teacher=None,
            classroom=_room_name(parts[-1]),
            group_number=group,
            class_name=parts[0].replace(" ", ""),
        )
    subject, group = _split_group(parts[0])
    if not subject:
        return None
    room = _room_name(parts[1]) if len(parts) > 1 and _ROOM_PREFIX_RE.match(parts[1]) else None
    teacher = _empty_to_none(parts[2]) if len(parts) > 2 else None
    if room is None and last_is_room:
        room = _room_name(parts[-1])
    return CellEntry(
        subject=subject,
        teacher=teacher,
        classroom=room,
        group_number=group,
    )


def _parse_block_lines(lines: list[str]) -> list[CellEntry]:
    entries: list[CellEntry] = []
    index = 0
    while index < len(lines):
        if _ROOM_PREFIX_RE.match(lines[index]):
            index += 1
            continue
        subject, group = _split_group(lines[index])
        index += 1
        room = None
        teacher = None
        if index < len(lines) and _ROOM_PREFIX_RE.match(lines[index]):
            room = _room_name(lines[index])
            index += 1
        if index < len(lines) and not _ROOM_PREFIX_RE.match(lines[index]):
            next_is_new_block = index + 1 < len(lines) and _ROOM_PREFIX_RE.match(
                lines[index + 1]
            )
            if not next_is_new_block:
                teacher = _empty_to_none(lines[index])
                index += 1
        if subject:
            entries.append(
                CellEntry(
                    subject=subject,
                    teacher=teacher,
                    classroom=room,
                    group_number=group,
                )
            )
    return entries


def _sheet_rows(ws) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_str(value) for value in row])
    while rows and all(not cell for cell in rows[-1]):
        rows.pop()
    return rows


def _title_meta(title: str) -> tuple[str | None, str | None]:
    rest = (title or "").strip()
    prefix = "расписание "
    if rest.casefold().startswith(prefix):
        rest = rest[len(prefix) :].strip()
    if _is_class_name(rest):
        return rest.replace(" ", ""), None
    if rest:
        return None, rest
    return None, None


def parse_schedule_workbook(path) -> list[ParsedSlot]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        slots: list[ParsedSlot] = []
        for sheet in workbook.worksheets:
            slots.extend(_parse_sheet(sheet.title, _sheet_rows(sheet)))
    finally:
        workbook.close()
    if not slots:
        raise ValueError(
            "В файле нет уроков в формате экспорта расписания "
            "(полное расписание, класс или учитель)"
        )
    return slots


def _parse_sheet(title: str, rows: list[list[str]]) -> list[ParsedSlot]:
    if not rows:
        return []
    first = rows[0][0] if rows[0] else ""
    if _day_number(first.split("\n", 1)[0].strip()):
        return _parse_shift_sheet(rows)
    if first.split("\n", 1)[0].strip().casefold() == "урок":
        return _parse_week_sheet(title, rows)
    return []


def _parse_shift_sheet(rows: list[list[str]]) -> list[ParsedSlot]:
    slots: list[ParsedSlot] = []
    current_day: int | None = None
    class_cols: list[tuple[int, str]] = []
    for row in rows:
        if not row:
            continue
        index_text = row[0]
        index_first = index_text.split("\n", 1)[0].strip()
        day = _day_number(index_first)
        if day is not None:
            current_day = day
            continue
        if index_first.casefold() == "урок":
            class_cols = []
            for col, value in enumerate(row[1:], start=1):
                if value:
                    class_cols.append((col, value.replace(" ", "")))
            continue
        if _is_skip_index(index_text):
            continue
        lesson = _lesson_number(index_text)
        if lesson is None or current_day is None or not class_cols:
            continue
        for col, class_name in class_cols:
            cell_text = row[col] if col < len(row) else ""
            for entry in parse_cell_entries(cell_text):
                slots.append(
                    ParsedSlot(
                        class_name=class_name,
                        day_of_week=current_day,
                        lesson_number=lesson,
                        subject=entry.subject,
                        teacher=entry.teacher,
                        classroom=entry.classroom,
                        group_number=entry.group_number,
                    )
                )
    return slots


def _parse_week_sheet(title: str, rows: list[list[str]]) -> list[ParsedSlot]:
    header = rows[0]
    day_cols: list[tuple[int, int]] = []
    for col, value in enumerate(header[1:], start=1):
        day = _day_number(value)
        if day is not None:
            day_cols.append((col, day))
    if not day_cols:
        return []
    title_class, title_teacher = _title_meta(title)
    slots: list[ParsedSlot] = []
    for row in rows[1:]:
        if not row:
            continue
        index_text = row[0]
        if _is_skip_index(index_text):
            continue
        lesson = _lesson_number(index_text)
        if lesson is None:
            continue
        for col, day in day_cols:
            cell_text = row[col] if col < len(row) else ""
            for entry in parse_cell_entries(cell_text):
                class_name = entry.class_name or title_class
                teacher = entry.teacher or title_teacher
                if not class_name:
                    continue
                slots.append(
                    ParsedSlot(
                        class_name=class_name,
                        day_of_week=day,
                        lesson_number=lesson,
                        subject=entry.subject,
                        teacher=teacher,
                        classroom=entry.classroom,
                        group_number=entry.group_number,
                    )
                )
    return slots


def import_schedule_from_excel(
    session: Session,
    school_id: int,
    path,
    *,
    replace: bool = False,
) -> ImportScheduleResultData:
    slots = parse_schedule_workbook(path)
    classes = list(
        session.scalars(
            select(SchoolClass).where(SchoolClass.school_id == school_id)
        ).all()
    )
    subjects = list(
        session.scalars(select(Subject).where(Subject.school_id == school_id)).all()
    )
    teachers = list(
        session.scalars(select(Teacher).where(Teacher.school_id == school_id)).all()
    )
    classrooms = list(
        session.scalars(
            select(Classroom).where(Classroom.school_id == school_id)
        ).all()
    )
    assignments = list(
        session.scalars(
            select(TeachingAssignment)
            .where(TeachingAssignment.school_id == school_id)
            .options(
                selectinload(TeachingAssignment.subject),
                selectinload(TeachingAssignment.teacher),
                selectinload(TeachingAssignment.school_class),
            )
        ).all()
    )

    class_by_name = {_norm_class(c.name): c for c in classes}
    subject_by_name = {s.display_name.casefold(): s for s in subjects}
    teacher_by_name = {normalize_person_name(t.full_name): t for t in teachers}
    classroom_by_display = {c.display_name.casefold(): c for c in classrooms}
    classroom_by_number = {c.number.casefold(): c for c in classrooms}

    assignments_by_class: dict[int, list[TeachingAssignment]] = {}
    for assignment in assignments:
        assignments_by_class.setdefault(assignment.class_id, []).append(assignment)

    warnings: list[str] = []
    unmatched = 0
    placements: list[Placement] = []
    seen: set[tuple[int, int, int, int]] = set()

    for slot in slots:
        school_class = class_by_name.get(_norm_class(slot.class_name))
        if school_class is None:
            unmatched += 1
            warnings.append(f"Класс «{slot.class_name}» не найден в справочнике")
            continue

        subject = subject_by_name.get(slot.subject.casefold())
        if subject is None:
            unmatched += 1
            warnings.append(
                f"{slot.class_name}, день {slot.day_of_week}, урок {slot.lesson_number}: "
                f"предмет «{slot.subject}» не найден"
            )
            continue

        teacher = None
        if slot.teacher:
            teacher = teacher_by_name.get(normalize_person_name(slot.teacher))
            if teacher is None:
                teacher = _match_truncated_teacher(slot.teacher, teachers)
            if teacher is None:
                unmatched += 1
                warnings.append(
                    f"{slot.class_name}, день {slot.day_of_week}, урок {slot.lesson_number}: "
                    f"учитель «{slot.teacher}» не найден"
                )
                continue

        assignment = _match_assignment(
            assignments_by_class.get(school_class.id, []),
            subject_id=subject.id,
            teacher_id=teacher.id if teacher is not None else None,
            teacher_required=bool(slot.teacher),
            group_number=slot.group_number,
        )
        if assignment is None:
            unmatched += 1
            teacher_label = slot.teacher or "без учителя"
            warnings.append(
                f"{slot.class_name}, день {slot.day_of_week}, урок {slot.lesson_number}: "
                f"нет назначения «{slot.subject}» / {teacher_label}"
            )
            continue

        classroom_id = None
        if slot.classroom:
            room = classroom_by_display.get(slot.classroom.casefold())
            if room is None:
                room = classroom_by_number.get(slot.classroom.casefold())
            if room is None:
                warnings.append(
                    f"{slot.class_name}, день {slot.day_of_week}, урок {slot.lesson_number}: "
                    f"кабинет «{slot.classroom}» не найден — урок без кабинета"
                )
            else:
                classroom_id = room.id

        key = (
            school_class.id,
            slot.day_of_week,
            slot.lesson_number,
            assignment.id,
        )
        if key in seen:
            continue
        seen.add(key)
        placements.append(
            Placement(
                assignment_id=assignment.id,
                class_id=school_class.id,
                day_of_week=slot.day_of_week,
                lesson_number=slot.lesson_number,
                classroom_id=classroom_id,
            )
        )

    warnings = _unique_keep_order(warnings)
    schedule = ScheduleService(session, school_id)
    cleared = 0
    restore_class_ids = {p.class_id for p in placements}
    if replace and restore_class_ids:
        cleared = schedule.delete_cells(class_ids=list(restore_class_ids), commit=False)

    existing = {
        (cell.class_id, cell.day_of_week, cell.lesson_number, cell.assignment_id)
        for cell in session.scalars(
            select(ScheduleCell).where(ScheduleCell.school_id == school_id)
        ).all()
    }
    skipped_existing = 0
    to_insert: list[Placement] = []
    for placement in placements:
        key = (
            placement.class_id,
            placement.day_of_week,
            placement.lesson_number,
            placement.assignment_id,
        )
        if key in existing:
            skipped_existing += 1
            continue
        to_insert.append(placement)
        existing.add(key)

    placed = 0
    if to_insert:
        placed = schedule.apply_placements(to_insert, validate=False, commit=True)
    elif cleared:
        session.commit()

    message = f"Поставлено уроков: {placed}"
    if cleared:
        message += f", очищено предыдущих: {cleared}"
    if skipped_existing:
        message += f", уже стояли: {skipped_existing}"
    if unmatched:
        message += f", не распознано: {unmatched}"
    return ImportScheduleResultData(
        placed=placed,
        skipped_existing=skipped_existing,
        unmatched=unmatched,
        cleared=cleared,
        warnings=warnings,
        message=message,
    )


def _match_truncated_teacher(name: str, teachers: list[Teacher]) -> Teacher | None:
    key = normalize_person_name(name)
    if not key:
        return None
    hits = [t for t in teachers if normalize_person_name(t.full_name).startswith(key)]
    if len(hits) == 1:
        return hits[0]
    return None


def _match_assignment(
    candidates: list[TeachingAssignment],
    *,
    subject_id: int,
    teacher_id: int | None,
    teacher_required: bool,
    group_number: int | None,
) -> TeachingAssignment | None:
    matched = [a for a in candidates if a.subject_id == subject_id]
    if teacher_required:
        matched = [a for a in matched if a.teacher_id == teacher_id]
    elif teacher_id is None:
        without_teacher = [a for a in matched if a.teacher_id is None]
        if without_teacher:
            matched = without_teacher
    if group_number is not None:
        grouped = [a for a in matched if a.group_number == group_number]
        if grouped:
            matched = grouped
    if len(matched) == 1:
        return matched[0]
    if len(matched) > 1 and group_number is None:
        ungrouped = [a for a in matched if a.group_number is None]
        if len(ungrouped) == 1:
            return ungrouped[0]
    return None


def _unique_keep_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        if item in seen:
            continue
        seen.add(item)
        result.append(item)
    return result
